#!/usr/bin/env python3
"""
Job Monitor for Raaga Sindhu
Profile: MS Data Analytics Engineering | 3+ years | Python, SQL, Snowflake, BigQuery,
         Tableau, Power BI, Scikit-learn, ETL/ELT, Airflow, Dask, Vertex AI

Target roles (tailored to her resume):
  - Data Analyst / Senior Data Analyst
  - Analytics Engineer
  - BI Analyst / BI Engineer
  - Data Scientist (Applied / Research)
  - Research Analyst / Quantitative Analyst
  - Product Analyst / Product Data Scientist
  - Decision Scientist
  - Workforce / People Analytics Analyst
  - Applied ML / ML Analyst

Sources:
  - RemoteOK       (free API, no key needed)
  - Arbeitnow      (free API, no key needed)
  - The Muse       (free API, no key needed)
  - Himalayas      (free API, no key needed)
  - Greenhouse     (public board API — major tech companies)
  - Lever          (public posting API — major tech companies)
  - jobright.ai    (scrape public listings)
  - Indeed         (RSS feed scrape)

Outputs:
  - jobs.json      (raw data for dashboard JS)
  - index.html     (full interactive dashboard — served via GitHub Pages)
  - job_monitor.xlsx (spreadsheet download)
"""

import requests
import json
import re
import time
import os
from datetime import datetime, timezone
from pathlib import Path
from bs4 import BeautifulSoup
import pandas as pd

# ─── Config — Tailored to Raaga's profile ───────────────────────────────────────
TARGET_TITLES = [
    # Core strength — Data Analyst
    "data analyst", "senior data analyst", "data analyst ii", "data analyst iii",
    # Analytics Engineering
    "analytics engineer", "analytics engineering", "analytic engineer",
    # BI
    "business intelligence", "bi analyst", "bi engineer", "bi developer",
    "business analyst",
    # Data Science
    "data scientist", "data science",
    # ML (NASA pipeline experience)
    "machine learning", "ml engineer", "ml analyst", "applied scientist",
    # Research / Quant
    "research analyst", "quantitative analyst", "quant analyst", "decision scientist",
    # Product analytics
    "product analyst", "product data", "growth analyst", "marketing analyst",
    "marketing data", "insights analyst",
    # People / Workforce Analytics
    "people analyst", "people analytics", "workforce analyst", "workforce analytics",
    "hr analyst", "hr analytics",
    # General data roles
    "data science analyst", "reporting analyst", "operations analyst",
]

EXCLUDE_TITLES = [
    "director", "vp ", "vice president", "head of", "intern", "co-op",
    "data entry", "data governance lead", "chief "
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json",
}

TIMESTAMP = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
ROOT = Path(__file__).parent.parent

# ─── H1B Sponsor List (DOL LCA disclosure data — top tech/data employers) ────────
# Source: USCIS/DOL public H1B filing records, curated for data/analytics roles
H1B_SPONSORS = {
    # Big Tech (own ATS systems — monitored via job boards)
    "amazon", "google", "alphabet", "microsoft", "meta", "apple", "netflix",
    "tesla", "uber", "lyft", "airbnb", "twitter", "x corp",
    # Cloud / Data Platforms
    "databricks", "snowflake", "confluent", "mongodb", "elastic", "datadog",
    "dbt labs", "fivetran", "palantir", "splunk", "tableau", "looker",
    "informatica", "talend", "alteryx", "qlik", "microstrategy",
    # Enterprise Software
    "salesforce", "oracle", "sap", "adobe", "cisco", "intel", "ibm",
    "qualcomm", "nvidia", "amd", "broadcom", "vmware", "servicenow",
    "workday", "veeva", "zendesk", "hubspot", "twilio",
    # Consulting / Big 4 (heavy H1B sponsors)
    "deloitte", "accenture", "pwc", "ernst & young", "ey", "kpmg",
    "mckinsey", "bcg", "bain", "boston consulting", "capgemini",
    "cognizant", "infosys", "tcs", "tata consultancy", "wipro",
    "hcl", "hcl technologies", "tech mahindra", "mphasis", "hexaware",
    "ltimindtree", "persistent systems",
    # Finance / Banking
    "jpmorgan", "jp morgan", "goldman sachs", "morgan stanley",
    "bank of america", "wells fargo", "citibank", "citi", "capital one",
    "american express", "amex", "visa", "mastercard", "paypal",
    "stripe", "square", "block", "robinhood", "coinbase",
    "blackrock", "vanguard", "fidelity", "charles schwab",
    # Healthcare / Biotech
    "unitedhealth", "anthem", "aetna", "humana", "cigna",
    "johnson & johnson", "pfizer", "abbvie", "amgen", "genentech",
    "roche", "novartis", "merck", "eli lilly", "bristol myers",
    "illumina", "10x genomics", "moderna",
    # Retail / E-commerce
    "walmart", "target", "costco", "kroger",
    "wayfair", "chewy", "etsy", "ebay",
    # Media / Entertainment
    "disney", "warner", "comcast", "nbcuniversal",
    "spotify", "soundcloud", "twitch",
    # Telecom
    "at&t", "verizon", "t-mobile",
    # Automotive / Manufacturing
    "ford", "general motors", "gm", "boeing", "lockheed",
    # Startups / SaaS (known H1B sponsors)
    "stripe", "plaid", "chime", "brex", "ramp", "rippling", "gusto",
    "amplitude", "mixpanel", "segment", "braze", "klaviyo",
    "figma", "notion", "airtable", "asana", "monday",
    "grammarly", "canva", "miro", "retool",
    "scale ai", "scale-ai", "cohere", "anthropic", "openai",
    "hugging face", "weights & biases", "wandb",
    "reddit", "discord", "pinterest", "quora",
    "doordash", "instacart", "gopuff",
    "opendoor", "compass", "zillow", "redfin",
    "duolingo", "coursera", "udemy", "chegg",
    "zendesk", "freshworks", "sprinklr",
    "pagerduty", "hashicorp", "cloudflare", "fastly",
    "dataiku", "domino data lab", "c3 ai",
    "veritone", "appen", "labelbox",
}

def is_h1b_sponsor(company: str) -> bool:
    """Check if a company is a known H1B sponsor."""
    c = company.lower().strip()
    return any(sponsor in c or c in sponsor for sponsor in H1B_SPONSORS)

# ─── Workday companies (POST API — no key needed) ────────────────────────────────
# Format: (workday_slug, instance_num, career_site_name, display_name)
WORKDAY_COMPANIES = [
    # Enterprise / Consulting (heavy H1B sponsors)
    ("salesforce",      "12", "External_Career_Website",        "Salesforce"),
    ("adobe",           "5",  "Experienced_Hire",               "Adobe"),
    ("cisco",           "5",  "Cisco",                          "Cisco"),
    ("intel",           "1",  "External",                       "Intel"),
    ("oracle",          "1",  "External",                       "Oracle"),
    ("capitalone",      "1",  "Capital_One",                    "Capital One"),
    ("goldmansachs",    "1",  "External_Career_Site",           "Goldman Sachs"),
    ("jpmorganchase",   "5",  "External",                       "JPMorgan Chase"),
    ("deloitteus",      "1",  "Experienced_Hire",               "Deloitte"),
    ("mckinsey",        "12", "McKinsey_Global",                "McKinsey"),
    ("walmart",         "5",  "WorkingAtWalmart",               "Walmart"),
    ("target",          "5",  "T-Target",                       "Target"),
    ("accenture",       "103","Accenture-External-Career-Site", "Accenture"),
    ("paypal",          "5",  "External",                       "PayPal"),
    ("qualcomm",        "5",  "External",                       "Qualcomm"),
    ("nvidia",          "1",  "External",                       "NVIDIA"),
    ("doordash",        "5",  "External",                       "DoorDash"),
    ("wayfair",         "5",  "Wayfair_Careers",                "Wayfair"),
    ("duolingo",        "5",  "External",                       "Duolingo"),
    ("spotify",         "5",  "External",                       "Spotify"),
    ("pagerduty",       "5",  "External",                       "PagerDuty"),
]

# ─── Ashby companies (public API — many funded startups) ─────────────────────────
ASHBY_COMPANIES = [
    "openai", "perplexity", "mistral", "together-ai", "anyscale",
    "modal", "replicate", "baseten", "bentoml",
    "dbtlabs", "lightdash", "evidence", "rill",
    "motherduck", "turntable", "recap",
    "linear", "retool", "airplane", "superblocks",
    "hex", "deepnote", "marimo",
    "census", "hightouch", "polytomic",
    "synq", "atlan", "secoda", "metaphor",
    "chalk", "tecton", "fennel",
    "continual", "superwise", "arize",
]

# ─── Greenhouse companies — strong data/analytics hiring ────────────────────────
# Focused on companies that hire data analysts, analytics engineers, BI, data scientists
GREENHOUSE_COMPANIES = [
    # Big tech / cloud data platforms
    "databricks", "snowflake", "confluent", "mongodb", "elastic", "datadog",
    "dbt-labs", "fivetran", "hightouch", "astronomer", "prefect",
    # AI/ML companies (Raaga's ML background)
    "anthropic", "openai", "cohere", "scale-ai", "weights-and-biases",
    "huggingface", "labelbox",
    # Product / consumer tech
    "airbnb", "stripe", "figma", "notion", "airtable", "zapier",
    "webflow", "asana", "monday", "clickup",
    # Analytics / BI focused
    "looker", "mode", "hex", "metabase", "preset",
    # Ed-tech / workforce (matches WGU background)
    "coursera", "udemy", "chegg", "canvas",
    # Healthcare / fintech (strong analytics demand)
    "robinhood", "brex", "plaid", "chime", "oscar",
]

# ─── Lever companies — strong analytics/data hiring ─────────────────────────────
LEVER_COMPANIES = [
    # Consumer / marketplace
    "lyft", "pinterest", "reddit", "discord", "etsy",
    # SaaS / analytics tools
    "amplitude", "mixpanel", "segment", "heap", "pendo",
    "braze", "klaviyo", "attentive", "iterable",
    # Productivity / collab
    "grammarly", "canva", "miro", "retool", "coda",
    # Data infrastructure
    "census-data", "lightdash", "streamlit",
    # Fintech
    "mercury", "ramp", "carta", "rippling", "gusto",
    # HR / workforce analytics (matches WGU)
    "lattice", "culture-amp", "leapsome", "workday",
]

# ─── Helpers ─────────────────────────────────────────────────────────────────────
def is_relevant(title: str, desc: str = "") -> bool:
    t = title.lower()
    if not any(kw in t for kw in TARGET_TITLES):
        return False
    if any(kw in t for kw in EXCLUDE_TITLES):
        return False
    return True

def clean_html(text: str, max_len: int = 400) -> str:
    if not text:
        return ""
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'&nbsp;', ' ', text)
    text = re.sub(r'&amp;', '&', text)
    text = re.sub(r'&#\d+;', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text[:max_len]

def safe_get(url, extra_headers=None, timeout=25, **kwargs) -> requests.Response | None:
    """Make a GET request, merging any extra headers with defaults."""
    try:
        hdrs = {**HEADERS, **(extra_headers or {})}
        r = requests.get(url, headers=hdrs, timeout=timeout, **kwargs)
        if r.status_code == 200:
            return r
    except Exception as e:
        print(f"    ⚠ Request failed for {url[:60]}: {e}")
    return None

# ─── RemoteOK ────────────────────────────────────────────────────────────────────
def fetch_remoteok() -> list[dict]:
    jobs, seen = [], set()
    tags = ["data-science", "machine-learning", "data-analyst", "ai", "nlp"]
    for tag in tags:
        r = safe_get(f"https://remoteok.com/api?tags={tag}")
        if not r:
            continue
        for item in r.json():
            if not isinstance(item, dict) or "position" not in item:
                continue
            job_id = str(item.get("id", ""))
            if job_id in seen:
                continue
            seen.add(job_id)
            title = item.get("position", "")
            if not is_relevant(title):
                continue
            jobs.append({
                "id": f"remoteok-{job_id}",
                "title": title,
                "company": item.get("company", "N/A"),
                "location": "Remote",
                "url": item.get("url", ""),
                "posted": (item.get("date", "") or "")[:10],
                "source": "RemoteOK",
                "tags": item.get("tags", [])[:6],
                "description": clean_html(item.get("description", "")),
                "salary": str(item.get("salary_min", "") or ""),
            })
        time.sleep(0.5)
    print(f"  RemoteOK: {len(jobs)} jobs")
    return jobs

# ─── Arbeitnow ───────────────────────────────────────────────────────────────────
def fetch_arbeitnow() -> list[dict]:
    jobs = []
    for page in range(1, 5):
        r = safe_get(f"https://www.arbeitnow.com/api/job-board-api?page={page}")
        if not r:
            break
        data = r.json().get("data", [])
        if not data:
            break
        for item in data:
            title = item.get("title", "")
            if not is_relevant(title):
                continue
            loc = item.get("location", "")
            if item.get("remote"):
                loc = "Remote" if not loc else f"{loc} (Remote)"
            jobs.append({
                "id": f"arbeitnow-{item.get('slug', title[:20])}",
                "title": title,
                "company": item.get("company_name", "N/A"),
                "location": loc or "N/A",
                "url": item.get("url", ""),
                "posted": (item.get("created_at", "") or "")[:10],
                "source": "Arbeitnow",
                "tags": item.get("tags", [])[:6],
                "description": clean_html(item.get("description", "")),
                "salary": "",
            })
        time.sleep(0.3)
    print(f"  Arbeitnow: {len(jobs)} jobs")
    return jobs

# ─── The Muse ────────────────────────────────────────────────────────────────────
def fetch_themuse() -> list[dict]:
    jobs, seen = [], set()
    categories = ["Data+Science", "Data+Analytics", "Software+Engineer"]
    for cat in categories:
        for page in range(1, 4):
            r = safe_get(f"https://www.themuse.com/api/public/jobs?category={cat}&level=Mid+Level&page={page}&per_page=20")
            if not r:
                continue
            for item in r.json().get("results", []):
                title = item.get("name", "")
                job_id = str(item.get("id", ""))
                if job_id in seen or not is_relevant(title):
                    continue
                seen.add(job_id)
                locs = item.get("locations", [])
                loc = locs[0].get("name", "N/A") if locs else "N/A"
                jobs.append({
                    "id": f"muse-{job_id}",
                    "title": title,
                    "company": item.get("company", {}).get("name", "N/A"),
                    "location": loc,
                    "url": item.get("refs", {}).get("landing_page", ""),
                    "posted": (item.get("publication_date", "") or "")[:10],
                    "source": "The Muse",
                    "tags": [c.get("name", "") for c in item.get("categories", [])[:4]],
                    "description": clean_html(item.get("contents", "")),
                    "salary": "",
                })
        time.sleep(0.3)
    print(f"  The Muse: {len(jobs)} jobs")
    return jobs

# ─── Himalayas ───────────────────────────────────────────────────────────────────
def fetch_himalayas() -> list[dict]:
    jobs, seen = [], set()
    queries = ["data-scientist", "machine-learning-engineer", "data-analyst", "ai-engineer"]
    for q in queries:
        r = safe_get(f"https://himalayas.app/jobs/api?q={q}&limit=25")
        if not r:
            continue
        data = r.json()
        items = data if isinstance(data, list) else data.get("jobs", [])
        for item in items:
            title = item.get("title", "")
            job_id = str(item.get("id", title + item.get("companyName", "")))
            if job_id in seen or not is_relevant(title):
                continue
            seen.add(job_id)
            locs = item.get("locationRestrictions", ["Worldwide"])
            jobs.append({
                "id": f"himalayas-{job_id[:30]}",
                "title": title,
                "company": item.get("companyName", "N/A"),
                "location": locs[0] if locs else "Remote",
                "url": item.get("applyUrl", item.get("url", "")),
                "posted": (item.get("createdAt", "") or "")[:10],
                "source": "Himalayas",
                "tags": item.get("skills", [])[:6],
                "description": clean_html(item.get("description", "")),
                "salary": str(item.get("salaryMin", "") or ""),
            })
        time.sleep(0.3)
    print(f"  Himalayas: {len(jobs)} jobs")
    return jobs

# ─── Greenhouse ──────────────────────────────────────────────────────────────────
def fetch_greenhouse() -> list[dict]:
    jobs = []
    for company in GREENHOUSE_COMPANIES:
        r = safe_get(f"https://boards-api.greenhouse.io/v1/boards/{company}/jobs?content=true")
        if not r:
            continue
        try:
            data = r.json().get("jobs", [])
        except Exception:
            continue
        for item in data:
            title = item.get("title", "")
            if not is_relevant(title):
                continue
            loc = item.get("location", {}).get("name", "N/A")
            jobs.append({
                "id": f"greenhouse-{item.get('id', '')}",
                "title": title,
                "company": company.replace("-", " ").title(),
                "location": loc,
                "url": item.get("absolute_url", ""),
                "posted": (item.get("updated_at", "") or "")[:10],
                "source": "Greenhouse",
                "tags": [d.get("value", "") for d in item.get("departments", [])[:3]],
                "description": clean_html(item.get("content", "")),
                "salary": "",
            })
        time.sleep(0.2)
    print(f"  Greenhouse: {len(jobs)} jobs")
    return jobs

# ─── Lever ───────────────────────────────────────────────────────────────────────
def fetch_lever() -> list[dict]:
    jobs = []
    for company in LEVER_COMPANIES:
        r = safe_get(f"https://api.lever.co/v0/postings/{company}?mode=json&limit=50", timeout=10)
        if not r:
            continue
        try:
            data = r.json()
        except Exception:
            continue
        if not isinstance(data, list):
            continue
        for item in data:
            title = item.get("text", "")
            if not is_relevant(title):
                continue
            categories = item.get("categories", {})
            loc = categories.get("location", "N/A")
            desc_lists = item.get("descriptionPlain", "") or clean_html(item.get("description", ""))
            jobs.append({
                "id": f"lever-{item.get('id', '')}",
                "title": title,
                "company": company.replace("-", " ").title(),
                "location": loc,
                "url": item.get("hostedUrl", item.get("applyUrl", "")),
                "posted": datetime.fromtimestamp(
                    item.get("createdAt", 0) / 1000, tz=timezone.utc
                ).strftime("%Y-%m-%d") if item.get("createdAt") else "",
                "source": "Lever",
                "tags": [t for t in [categories.get("team", ""), categories.get("commitment", "")] if t],
                "description": desc_lists[:400],
                "salary": "",
            })
        time.sleep(0.2)
    print(f"  Lever: {len(jobs)} jobs")
    return jobs

# ─── Indeed RSS ──────────────────────────────────────────────────────────────────
def fetch_indeed_rss() -> list[dict]:
    jobs = []
    queries = [
        ("data+scientist", "mid level"),
        ("machine+learning+engineer", "mid level"),
        ("data+analyst", "mid level"),
    ]
    for q, level in queries:
        url = f"https://www.indeed.com/rss?q={q}&sort=date&limit=25"
        r = safe_get(url)
        if not r:
            continue
        try:
            soup = BeautifulSoup(r.text, "lxml-xml")
            for item in soup.find_all("item"):
                title = item.find("title")
                title = title.text if title else ""
                if not is_relevant(title):
                    continue
                link = item.find("link")
                link = link.text if link else ""
                company_tag = item.find("source")
                company = company_tag.text if company_tag else "N/A"
                loc_tag = item.find("indeedapply:jobtitle") or item.find("location")
                loc = loc_tag.text if loc_tag else "N/A"
                desc_tag = item.find("description")
                desc = clean_html(desc_tag.text if desc_tag else "")
                pub_tag = item.find("pubDate")
                pub = pub_tag.text[:10] if pub_tag else ""
                jobs.append({
                    "id": f"indeed-{hash(title+company)}",
                    "title": title,
                    "company": company,
                    "location": loc,
                    "url": link,
                    "posted": pub,
                    "source": "Indeed",
                    "tags": [],
                    "description": desc,
                    "salary": "",
                })
        except Exception as e:
            print(f"    ⚠ Indeed RSS parse error: {e}")
        time.sleep(0.5)
    print(f"  Indeed RSS: {len(jobs)} jobs")
    return jobs

# ─── Jobright.ai scrape ──────────────────────────────────────────────────────────
def fetch_jobright() -> list[dict]:
    jobs = []
    queries = ["data-scientist", "machine-learning-engineer", "data-analyst"]
    for q in queries:
        url = f"https://jobright.ai/jobs/{q}"
        r = safe_get(url, extra_headers={"Accept": "text/html"})
        if not r:
            continue
        try:
            soup = BeautifulSoup(r.text, "html.parser")
            # Find job cards — jobright uses data-testid or class-based selectors
            cards = soup.find_all("div", attrs={"data-testid": re.compile("job-card|JobCard", re.I)})
            if not cards:
                cards = soup.find_all("li", class_=re.compile("job", re.I))
            for card in cards[:15]:
                title_el = card.find(["h2", "h3", "a"], class_=re.compile("title|name", re.I))
                if not title_el:
                    continue
                title = title_el.get_text(strip=True)
                if not is_relevant(title):
                    continue
                company_el = card.find(class_=re.compile("company|employer", re.I))
                company = company_el.get_text(strip=True) if company_el else "N/A"
                loc_el = card.find(class_=re.compile("location|loc", re.I))
                loc = loc_el.get_text(strip=True) if loc_el else "N/A"
                link_el = card.find("a", href=True)
                link = "https://jobright.ai" + link_el["href"] if link_el and link_el["href"].startswith("/") else (link_el["href"] if link_el else url)
                jobs.append({
                    "id": f"jobright-{hash(title+company)}",
                    "title": title,
                    "company": company,
                    "location": loc,
                    "url": link,
                    "posted": "",
                    "source": "Jobright.ai",
                    "tags": [],
                    "description": "",
                    "salary": "",
                })
        except Exception as e:
            print(f"    ⚠ Jobright parse error: {e}")
        time.sleep(0.5)
    print(f"  Jobright.ai: {len(jobs)} jobs")
    return jobs

# ─── Workday (POST API — no auth needed) ─────────────────────────────────────────
def fetch_workday() -> list[dict]:
    jobs, seen = [], set()
    search_terms = ["data analyst", "analytics engineer", "data scientist",
                    "machine learning", "business intelligence", "product analyst"]
    for slug, instance, site, display in WORKDAY_COMPANIES:
        base = f"https://{slug}.wd{instance}.myworkdayjobs.com/wday/cxs/{slug}/{site}/jobs"
        for term in search_terms:
            try:
                payload = {"appliedFacets": {}, "limit": 20, "offset": 0, "searchText": term}
                r = requests.post(base, json=payload,
                                  headers={**HEADERS, "Content-Type": "application/json"},
                                  timeout=15)
                if r.status_code != 200:
                    continue
                data = r.json()
                for item in data.get("jobPostings", []):
                    title = item.get("title", "")
                    job_id = item.get("externalPath", title)
                    if job_id in seen or not is_relevant(title):
                        continue
                    seen.add(job_id)
                    posted_on = item.get("postedOn", "")
                    loc = item.get("locationsText", "N/A")
                    url = f"https://{slug}.wd{instance}.myworkdayjobs.com/en-US/{site}{item.get('externalPath','')}"
                    jobs.append({
                        "id": f"workday-{slug}-{job_id[-20:]}",
                        "title": title,
                        "company": display,
                        "location": loc,
                        "url": url,
                        "posted": posted_on[:10] if posted_on else "",
                        "source": "Workday",
                        "tags": [],
                        "description": item.get("jobDescription", {}).get("descriptor", "")[:300],
                        "salary": "",
                        "h1b": is_h1b_sponsor(display),
                    })
                time.sleep(0.3)
            except Exception as e:
                print(f"    ⚠ Workday {display} '{term}': {e}")
        time.sleep(0.2)
    print(f"  Workday: {len(jobs)} jobs")
    return jobs

# ─── Ashby (funded startups — many sponsor H1B) ──────────────────────────────────
def fetch_ashby() -> list[dict]:
    jobs, seen = [], set()
    for company in ASHBY_COMPANIES:
        r = safe_get(f"https://api.ashbyhq.com/posting-public/job-board/{company}")
        if not r:
            continue
        try:
            data = r.json()
            for item in data.get("jobPostings", []):
                title = item.get("title", "")
                job_id = item.get("id", "")
                if job_id in seen or not is_relevant(title):
                    continue
                seen.add(job_id)
                loc_list = item.get("locationName", "") or item.get("isRemote", "")
                loc = loc_list if isinstance(loc_list, str) else ("Remote" if loc_list else "N/A")
                display = item.get("organizationName", company.title())
                jobs.append({
                    "id": f"ashby-{job_id}",
                    "title": title,
                    "company": display,
                    "location": loc,
                    "url": item.get("jobUrl", f"https://jobs.ashbyhq.com/{company}"),
                    "posted": (item.get("publishedDate", "") or "")[:10],
                    "source": "Ashby",
                    "tags": [item.get("departmentName", "")] if item.get("departmentName") else [],
                    "description": clean_html(item.get("descriptionHtml", ""))[:300],
                    "salary": "",
                    "h1b": is_h1b_sponsor(display),
                })
        except Exception as e:
            print(f"    ⚠ Ashby {company}: {e}")
        time.sleep(0.2)
    print(f"  Ashby: {len(jobs)} jobs")
    return jobs

# ─── Adzuna (aggregates LinkedIn, Indeed, Glassdoor & more) ─────────────────────
def fetch_adzuna() -> list[dict]:
    """
    Adzuna has a free public API (no key for basic use via their job search endpoint).
    Aggregates jobs from LinkedIn, Indeed, Glassdoor, and 100+ job boards.
    """
    jobs, seen = [], set()
    queries = [
        "data analyst", "analytics engineer", "data scientist",
        "machine learning engineer", "business intelligence analyst",
        "product analyst", "people analytics"
    ]
    for q in queries:
        q_enc = q.replace(" ", "%20")
        # Try US jobs (country=us)
        url = f"https://api.adzuna.com/v1/api/jobs/us/search/1?app_id=&app_key=&results_per_page=20&what={q_enc}&content-type=application/json"
        r = safe_get(url)
        if not r:
            # fallback: public search endpoint
            url2 = f"https://api.adzuna.com/v1/api/jobs/us/search/1?results_per_page=20&what={q_enc}"
            r = safe_get(url2)
        if not r:
            continue
        try:
            for item in r.json().get("results", []):
                title = item.get("title", "")
                job_id = str(item.get("id", ""))
                if job_id in seen or not is_relevant(title):
                    continue
                seen.add(job_id)
                loc = item.get("location", {}).get("display_name", "N/A")
                company = item.get("company", {}).get("display_name", "N/A")
                jobs.append({
                    "id": f"adzuna-{job_id}",
                    "title": title,
                    "company": company,
                    "location": loc,
                    "url": item.get("redirect_url", ""),
                    "posted": (item.get("created", "") or "")[:10],
                    "source": "Adzuna",
                    "tags": item.get("category", {}).get("label", "").split("/")[:4],
                    "description": clean_html(item.get("description", "")),
                    "salary": str(int(item.get("salary_min", 0))) if item.get("salary_min") else "",
                })
        except Exception as e:
            print(f"    ⚠ Adzuna parse error: {e}")
        time.sleep(0.3)
    print(f"  Adzuna: {len(jobs)} jobs")
    return jobs

# ─── LinkedIn Jobs via public feed ───────────────────────────────────────────────
def fetch_linkedin_feed() -> list[dict]:
    """
    LinkedIn public job search (no login, no API key).
    Uses their public job listing URLs with JSON-LD structured data.
    """
    jobs, seen = [], set()
    searches = [
        ("data analyst", "2"),
        ("analytics engineer", "2"),
        ("data scientist", "2"),
        ("machine learning engineer", "2"),
        ("business intelligence analyst", "2"),
        ("product analyst", "2"),
    ]
    for keywords, exp_level in searches:
        kw_enc = keywords.replace(" ", "%20")
        # f_E=2 = mid-senior level, f_TPR=r604800 = past week
        url = f"https://www.linkedin.com/jobs-guest/jobs/api/seeMoreJobPostings/search?keywords={kw_enc}&f_E={exp_level}&f_TPR=r604800&start=0"
        r = safe_get(url, extra_headers={"Accept": "text/html,application/xhtml+xml"})
        if not r:
            continue
        try:
            soup = BeautifulSoup(r.text, "html.parser")
            cards = soup.find_all("li")
            for card in cards[:10]:
                title_el = card.find("h3")
                title = title_el.get_text(strip=True) if title_el else ""
                if not title or not is_relevant(title):
                    continue
                company_el = card.find("h4")
                company = company_el.get_text(strip=True) if company_el else "N/A"
                loc_el = card.find("span", class_=re.compile("job-search-card__location", re.I))
                loc = loc_el.get_text(strip=True) if loc_el else "N/A"
                link_el = card.find("a", href=True)
                link = link_el["href"].split("?")[0] if link_el else ""
                job_id = link.split("/")[-1] if link else title + company
                if job_id in seen:
                    continue
                seen.add(job_id)
                time_el = card.find("time")
                posted = time_el.get("datetime", "")[:10] if time_el else ""
                jobs.append({
                    "id": f"linkedin-{job_id}",
                    "title": title,
                    "company": company,
                    "location": loc,
                    "url": link,
                    "posted": posted,
                    "source": "LinkedIn",
                    "tags": [],
                    "description": "",
                    "salary": "",
                })
        except Exception as e:
            print(f"    ⚠ LinkedIn parse error for '{keywords}': {e}")
        time.sleep(1)
    print(f"  LinkedIn: {len(jobs)} jobs")
    return jobs

# ─── Deduplicate ─────────────────────────────────────────────────────────────────
def deduplicate(jobs: list[dict]) -> list[dict]:
    seen, unique = set(), []
    for job in jobs:
        key = (job["title"].lower().strip()[:50], job["company"].lower().strip()[:30])
        if key not in seen:
            seen.add(key)
            unique.append(job)
    return unique

# ─── Excel ───────────────────────────────────────────────────────────────────────
def generate_excel(jobs: list[dict], path: Path):
    rows = []
    for j in jobs:
        rows.append({
            "Job Title": j["title"],
            "Company": j["company"],
            "H1B Sponsor": "✅ Yes" if j.get("h1b") else "❓ Unknown",
            "Location": j["location"],
            "Source": j["source"],
            "Posted": j["posted"],
            "Skills/Tags": ", ".join(j["tags"]) if isinstance(j["tags"], list) else j["tags"],
            "Salary": j["salary"],
            "Description": j["description"],
            "Apply URL": j["url"],
        })
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Job Listings")
        ws = writer.sheets["Job Listings"]
        widths = [35, 25, 20, 12, 12, 30, 15, 50, 50]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w
        from openpyxl.styles import PatternFill, Font, Alignment
        hfill = PatternFill("solid", fgColor="4F46E5")
        hfont = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        light = PatternFill("solid", fgColor="F0F0FF")
        for row in ws.iter_rows(min_row=2):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = light
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

# ─── HTML Dashboard ──────────────────────────────────────────────────────────────
def generate_html(jobs: list[dict]) -> str:
    jobs_json = json.dumps(jobs, ensure_ascii=False)
    source_counts = {}
    for j in jobs:
        s = j.get("source", "?")
        source_counts[s] = source_counts.get(s, 0) + 1

    source_colors = {
        "RemoteOK": "#4f46e5", "Arbeitnow": "#0891b2", "The Muse": "#059669",
        "Himalayas": "#d97706", "Greenhouse": "#16a34a", "Lever": "#dc2626",
        "Indeed": "#2563eb", "Jobright.ai": "#9333ea",
        "Adzuna": "#0f766e", "LinkedIn": "#0a66c2",
        "Workday": "#f59e0b", "Ashby": "#7c3aed"
    }

    # Pre-build dynamic HTML parts to avoid backslash-in-f-string errors (Python < 3.12)
    src_badges_parts = ['<span class="src-badge active" style="background:#334155" onclick="filterSource(this, \'all\')">All Sources</span>']
    for s, c in source_counts.items():
        color = source_colors.get(s, "#6b7280")
        badge = '<span class="src-badge" style="background:' + color + '" onclick="filterSource(this, \'' + s + '\')">' + s + ' (' + str(c) + ')</span>'
        src_badges_parts.append(badge)
    src_badges_html = "".join(src_badges_parts)

    total_jobs = len(jobs)
    total_sources = len(source_counts)
    source_colors_json = json.dumps(source_colors)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Raaga's Job Monitor — {TIMESTAMP}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f1f5f9;color:#1e293b}}
.header{{background:linear-gradient(135deg,#4f46e5,#7c3aed);color:#fff;padding:30px 36px}}
.header h1{{font-size:26px;font-weight:700;margin-bottom:6px}}
.header p{{opacity:.8;font-size:14px;margin-bottom:16px}}
.stats{{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:14px}}
.stat{{background:rgba(255,255,255,.15);border-radius:8px;padding:10px 18px;text-align:center}}
.stat-num{{font-size:22px;font-weight:700}}
.stat-label{{font-size:11px;opacity:.8;margin-top:2px}}
.src-badges{{display:flex;flex-wrap:wrap;gap:6px}}
.src-badge{{padding:4px 10px;border-radius:20px;font-size:11px;font-weight:600;color:#fff;cursor:pointer;transition:.2s;opacity:.75}}
.src-badge:hover,.src-badge.active{{opacity:1;transform:scale(1.05)}}
.controls{{background:#fff;padding:14px 36px;border-bottom:1px solid #e2e8f0;display:flex;gap:10px;flex-wrap:wrap;align-items:center;position:sticky;top:0;z-index:10;box-shadow:0 2px 8px rgba(0,0,0,.06)}}
input[type=text]{{padding:8px 14px;border:1px solid #cbd5e1;border-radius:8px;font-size:14px;width:240px;outline:none}}
input[type=text]:focus{{border-color:#4f46e5}}
.sort-select{{padding:8px 12px;border:1px solid #cbd5e1;border-radius:8px;font-size:13px;background:#fff;cursor:pointer}}
.count-label{{padding:12px 36px 0;color:#64748b;font-size:13px}}
.container{{max-width:1260px;margin:0 auto;padding:16px 20px 40px}}
.grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(360px,1fr));gap:18px}}
.card{{background:#fff;border-radius:12px;padding:20px;border:1px solid #e2e8f0;transition:.2s;display:flex;flex-direction:column;gap:10px}}
.card:hover{{transform:translateY(-2px);box-shadow:0 6px 20px rgba(0,0,0,.09)}}
.card-top{{display:flex;justify-content:space-between;align-items:flex-start;gap:10px}}
.job-title{{font-size:15px;font-weight:600;line-height:1.4}}
.job-title a{{color:#1e293b;text-decoration:none}}
.job-title a:hover{{color:#4f46e5}}
.company{{font-size:13px;color:#64748b;margin-top:3px}}
.src-pill{{padding:3px 9px;border-radius:20px;font-size:10px;font-weight:700;color:#fff;white-space:nowrap;flex-shrink:0}}
.desc{{font-size:12.5px;color:#475569;line-height:1.55}}
.meta{{display:flex;gap:8px;font-size:11.5px;color:#94a3b8;flex-wrap:wrap}}
.tags{{display:flex;gap:5px;flex-wrap:wrap}}
.tag{{background:#f1f5f9;color:#475569;padding:2px 8px;border-radius:5px;font-size:11px}}
.apply{{display:inline-block;background:#4f46e5;color:#fff;padding:7px 16px;border-radius:8px;text-decoration:none;font-size:12.5px;font-weight:600;margin-top:auto;transition:.15s;text-align:center}}
.apply:hover{{background:#4338ca}}
.no-results{{text-align:center;padding:70px 20px;color:#94a3b8;display:none}}
.h1b-badge{{background:#dcfce7;color:#15803d;padding:2px 8px;border-radius:5px;font-size:11px;font-weight:600}}
.h1b-btn{{padding:7px 14px;border:1.5px solid #15803d;border-radius:8px;background:white;color:#15803d;font-size:13px;font-weight:600;cursor:pointer;transition:.2s}}
.h1b-btn:hover{{background:#f0fdf4}}
footer{{text-align:center;padding:30px;color:#94a3b8;font-size:12px}}
@media(max-width:600px){{.grid{{grid-template-columns:1fr}}.header{{padding:20px}}.controls{{padding:10px 16px}}}}
</style>
</head>
<body>
<div class="header">
  <h1>🎯 Raaga's Job Monitor</h1>
  <p>Data Analyst · Analytics Engineer · BI Analyst · Data Scientist · Product Analyst · Decision Scientist &nbsp;|&nbsp; Mid-level · All locations</p>
  <div class="stats">
    <div class="stat"><div class="stat-num" id="totalJobs">{total_jobs}</div><div class="stat-label">Jobs Found</div></div>
    <div class="stat"><div class="stat-num">{total_sources}</div><div class="stat-label">Sources</div></div>
    <div class="stat"><div class="stat-num" id="visibleCount">{total_jobs}</div><div class="stat-label">Showing</div></div>
    <div class="stat"><div class="stat-num">{TIMESTAMP}</div><div class="stat-label">Last Refreshed</div></div>
  </div>
  <div class="src-badges">
    {src_badges_html}
  </div>
</div>

<div class="controls">
  <input type="text" id="searchBox" placeholder="🔍 Search title, company, skills..." oninput="render()">
  <select class="sort-select" id="sortBy" onchange="render()">
    <option value="posted">Sort: Most Recent</option>
    <option value="title">Sort: Title A–Z</option>
    <option value="company">Sort: Company A–Z</option>
    <option value="source">Sort: Source</option>
  </select>
  <button class="h1b-btn" id="h1bBtn" onclick="toggleH1B(this)">🟢 H1B Sponsors Only</button>
  <a href="job_monitor.xlsx" style="margin-left:auto;padding:7px 14px;background:#059669;color:#fff;border-radius:8px;text-decoration:none;font-size:13px;font-weight:600">⬇ Download Excel</a>
</div>

<p class="count-label" id="countLabel">Showing {total_jobs} jobs</p>
<div class="container">
  <div class="grid" id="jobGrid"></div>
  <div class="no-results" id="noResults">No jobs match your search. Try different keywords or clear filters.</div>
</div>
<footer>🤖 Auto-refreshes every 2 hours via GitHub Actions · {TIMESTAMP}</footer>

<script>
const ALL_JOBS = {jobs_json};
const SOURCE_COLORS = {source_colors_json};
let activeSource = 'all';
let h1bOnly = false;

function filterSource(el, src) {{
  activeSource = src;
  document.querySelectorAll('.src-badge').forEach(b => b.classList.remove('active'));
  el.classList.add('active');
  render();
}}

function toggleH1B(btn) {{
  h1bOnly = !h1bOnly;
  btn.style.background = h1bOnly ? '#15803d' : '';
  btn.style.color = h1bOnly ? '#fff' : '';
  render();
}}

function render() {{
  const q = document.getElementById('searchBox').value.toLowerCase();
  const sortBy = document.getElementById('sortBy').value;

  let filtered = ALL_JOBS.filter(j => {{
    const matchSrc = activeSource === 'all' || j.source === activeSource;
    const matchQ = !q || j.title.toLowerCase().includes(q) ||
      j.company.toLowerCase().includes(q) ||
      (j.tags || []).join(' ').toLowerCase().includes(q) ||
      j.description.toLowerCase().includes(q);
    const matchH1B = !h1bOnly || j.h1b === true;
    return matchSrc && matchQ && matchH1B;
  }});

  filtered.sort((a, b) => {{
    if (sortBy === 'posted') return (b.posted || '').localeCompare(a.posted || '');
    if (sortBy === 'title') return a.title.localeCompare(b.title);
    if (sortBy === 'company') return a.company.localeCompare(b.company);
    if (sortBy === 'source') return a.source.localeCompare(b.source);
    return 0;
  }});

  document.getElementById('countLabel').textContent = `Showing ${{filtered.length}} of ${{ALL_JOBS.length}} jobs`;
  document.getElementById('visibleCount').textContent = filtered.length;
  document.getElementById('noResults').style.display = filtered.length === 0 ? 'block' : 'none';

  const grid = document.getElementById('jobGrid');
  grid.innerHTML = filtered.map(j => {{
    const color = SOURCE_COLORS[j.source] || '#6b7280';
    const tags = (j.tags || []).slice(0,5).map(t => `<span class="tag">${{t}}</span>`).join('');
    const desc = j.description ? `<p class="desc">${{j.description.slice(0,220)}}${{j.description.length > 220 ? '…' : ''}}</p>` : '';
    const salary = j.salary ? `<span>💰 ${{j.salary}}</span>` : '';
    const posted = j.posted ? `<span>📅 ${{j.posted}}</span>` : '';
    const h1bBadge = j.h1b ? '<span class="h1b-badge">🟢 H1B Sponsor</span>' : '';
    return `
    <div class="card">
      <div class="card-top">
        <div>
          <div class="job-title"><a href="${{j.url}}" target="_blank">${{j.title}}</a></div>
          <div class="company">🏢 ${{j.company}} &nbsp;·&nbsp; 📍 ${{j.location}}</div>
        </div>
        <span class="src-pill" style="background:${{color}}">${{j.source}}</span>
      </div>
      ${{desc}}
      <div class="tags">${{tags}}</div>
      <div class="meta">${{h1bBadge}} ${{salary}} ${{posted}}</div>
      <a href="${{j.url}}" target="_blank" class="apply">Apply Now →</a>
    </div>`;
  }}).join('');
}}

render();
</script>
</body>
</html>"""

# ─── Main ────────────────────────────────────────────────────────────────────────
def main():
    print(f"🔍 Job Monitor — {TIMESTAMP}")
    print("=" * 50)

    all_jobs = []
    fetchers = [
        ("RemoteOK", fetch_remoteok),
        ("Arbeitnow", fetch_arbeitnow),
        ("The Muse", fetch_themuse),
        ("Himalayas", fetch_himalayas),
        ("Greenhouse", fetch_greenhouse),
        ("Lever", fetch_lever),
        ("Workday", fetch_workday),
        ("Ashby", fetch_ashby),
        ("Indeed RSS", fetch_indeed_rss),
        ("Jobright.ai", fetch_jobright),
        ("Adzuna", fetch_adzuna),
        ("LinkedIn", fetch_linkedin_feed),
    ]

    for name, fn in fetchers:
        print(f"→ {name}...")
        try:
            jobs = fn()
            all_jobs.extend(jobs)
        except Exception as e:
            print(f"  ⚠ {name} failed: {e}")

    all_jobs = deduplicate(all_jobs)
    # Tag every job with H1B sponsor status
    for job in all_jobs:
        if "h1b" not in job:
            job["h1b"] = is_h1b_sponsor(job.get("company", ""))
    all_jobs.sort(key=lambda x: x.get("posted", ""), reverse=True)
    h1b_count = sum(1 for j in all_jobs if j.get("h1b"))
    print(f"🟢 H1B sponsor jobs: {h1b_count} of {len(all_jobs)}")

    print(f"\n✅ Total unique jobs: {len(all_jobs)}")

    # Save jobs.json
    json_path = ROOT / "jobs.json"
    json_path.write_text(json.dumps(all_jobs, indent=2, ensure_ascii=False))
    print(f"📦 jobs.json saved ({len(all_jobs)} jobs)")

    # Save HTML
    html_path = ROOT / "index.html"
    html_path.write_text(generate_html(all_jobs), encoding="utf-8")
    print(f"🌐 index.html saved")

    # Save Excel
    xlsx_path = ROOT / "job_monitor.xlsx"
    generate_excel(all_jobs, xlsx_path)
    print(f"📊 job_monitor.xlsx saved")

    print(f"\n🎉 Done! Dashboard ready at your GitHub Pages URL.")

if __name__ == "__main__":
    main()
